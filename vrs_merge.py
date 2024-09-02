from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import shutil
from databases import Database
import asyncio
from datetime import datetime, timedelta

# 현재 스크립트의 디렉토리 경로를 가져옵니다.
current_dir = os.path.dirname(os.path.abspath(__file__))

# test.db 파일의 전체 경로를 생성합니다.
database_path = os.path.join(current_dir, "test.db")

# DATABASE_URL을 수정합니다.
DATABASE_URL = f"sqlite:///{database_path}"

database = Database(DATABASE_URL)

def is_merged_cell(sheet, row, col):
    """해당 셀이 병합된 셀의 첫 번째 셀인지 확인"""
    cell_coord = f"{get_column_letter(col)}{row}"
    for merged_range in sheet.merged_cells.ranges:
        if cell_coord in merged_range:
            start_coord = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
            return start_coord == cell_coord
    return True  # 병합되지 않은 셀은 True로 처리

# 엑셀 파일 업데이트 및 로그 기록 로직
def update_vrs_excel(date, team, worker, manager, equipment_id, items):
    # 날짜에서 년과 월을 추출
    year = date[:2]  # 예: '240826' -> '24'
    month = date[2:4]  # 예: '240826' -> '08'
    day = date[4:6]  # 예: '240826' -> '26'
    day = int(day)

    # 팀별 폴더 및 장비 호기별 폴더 설정
    target_directory = f"./excel/vrs/{team}/{equipment_id}"
    os.makedirs(target_directory, exist_ok=True)
    
    # 파일 이름 생성 및 경로 설정
    filename = f"VRS_{int(month)}월.xlsx"
    filepath = os.path.join(target_directory, filename)

    # 로그 파일 경로 설정
    log_filename = f"log_{int(month)}월.txt"
    log_filepath = os.path.join(target_directory, log_filename)

    # 교체일 로그 파일 경로 설정
    replace_log_filename = "replace_log.txt"
    replace_log_filepath = os.path.join(target_directory, replace_log_filename)

    # 로그 파일에서 중복 확인
    if os.path.exists(log_filepath):
        with open(log_filepath, 'r') as log_file:
            logs = log_file.readlines()
            for log in logs:
                if f"{year}/{int(month)}/{day}" in log:
                    print(f"이미 기록된 날짜: {year}/{int(month)}/{day}")
                    return  # 이미 기록된 날짜가 있으면 종료

    # 해당 월의 파일이 있는지 확인
    if os.path.exists(filepath):
        workbook = load_workbook(filename=filepath)
    else:
        # 파일이 없으면 기본 템플릿 파일(VRS.xlsx)을 복사하여 새로 생성
        shutil.copyfile("./checksheet/VRS.xlsx", filepath)
        workbook = load_workbook(filename=filepath)

    # 첫 번째 시트 업데이트
    sheet1 = workbook["DMTI-VRS-01-07 (설비점검)"]
    sheet1["AB3"] = f"{year}년 {int(month)}월"
    sheet1["F2"] = f"FQA ( {team}조)"
    sheet1["I1"] = f"VRS     {equipment_id}호기 설비 점검 Check Sheet"
    column_index = 7 + (day - 1)  # 첫 번째 시트의 1일은 G열
    rows_to_update_sheet1 = [7, 9, 11, 13, 15, 17]
    for row in rows_to_update_sheet1:
        if is_merged_cell(sheet1, row, column_index):
            sheet1.cell(row=row, column=column_index).value = 'O'

    # worker와 manager를 1번 시트의 22행, 23행에 입력
    sheet1.cell(row=22, column=column_index).value = worker
    sheet1.cell(row=23, column=column_index).value = manager

    # 두 번째 시트 업데이트
    sheet2 = workbook["DMTI-VRS-01-08 (청소점검)"]
    sheet2["AC3"] = f"{year}년 {int(month)}월"
    sheet2["F2"] = f"FQA ( {team}조)"
    sheet2["J1"] = f"VRS     {equipment_id}호기 설비 점검 Check Sheet"
    column_index = 8 + (day - 1)  # 두 번째 시트의 1일은 H열
    rows_to_update_sheet2 = [7, 9, 11, 13, 15, 17]
    for row in rows_to_update_sheet2:
        if is_merged_cell(sheet2, row, column_index):
            sheet2.cell(row=row, column=column_index).value = 'O'

    # worker와 manager를 2번 시트의 20행, 21행에 입력
    sheet2.cell(row=20, column=column_index).value = worker
    sheet2.cell(row=21, column=column_index).value = manager

    # 세 번째 시트 업데이트
    sheet3 = workbook["DMTI-VRS-01-08 (청소검증)"]
    sheet3["AB3"] = f"{year}년 {int(month)}월"
    sheet3["E2"] = f"FQA ( {team}조)"
    sheet3["I1"] = f"VRS     {equipment_id}호기 설비 점검 Check Sheet"
    column_index = 7 + (day - 1)  # 세 번째 시트의 1일은 G열
    rows_to_update_sheet3 = [7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29]
    for row in rows_to_update_sheet3:
        if is_merged_cell(sheet3, row, column_index):
            sheet3.cell(row=row, column=column_index).value = 'O'

    # 교체일 로직
    if os.path.exists(replace_log_filepath):
        with open(replace_log_filepath, 'r') as replace_log_file:
            replace_logs = replace_log_file.readlines()
            for line in replace_logs:
                if "G19" in line:
                    sheet1["G19"] = line.strip().replace("G19", "")
                if "G20" in line:
                    sheet1["G20"] = line.strip().replace("G20", "")
                if "G21" in line:
                    sheet1["G21"] = line.strip().replace("G21", "")
                if "H19" in line:
                    sheet2["H19"] = line.strip().replace("H19", "")

    if items.get(7) == 1:  # item_id 7이 1이면 교체일 갱신
        recent_date = f"최근 교체일: {date}"
        next_date = (datetime.strptime(date, "%y%m%d") + timedelta(days=90)).strftime("%y%m%d")
        next_date_str = f"다음 교체일: {next_date}"
        sheet1["G19"] = f"{recent_date}   {next_date_str}"
        sheet2["H19"] = f"{recent_date}   {next_date_str}"
        with open(replace_log_filepath, 'a') as replace_log_file: 
            replace_log_file.write(f"G19 {recent_date}   {next_date_str}\n")
            replace_log_file.write(f"H19 {recent_date}   {next_date_str}\n")

    if items.get(8) == 1:  # item_id 8이 1이면 교체일 갱신
        recent_date = f"최근 교체일: {date}"
        next_date = (datetime.strptime(date, "%y%m%d") + timedelta(days=365)).strftime("%y%m%d")
        next_date_str = f"다음 교체일: {next_date}"
        sheet1["G20"] = f"{recent_date}   {next_date_str}"
        with open(replace_log_filepath, 'a') as replace_log_file:
            replace_log_file.write(f"G20 {recent_date}   {next_date_str}\n")

    if items.get(9) == 1:  # item_id 9가 1이면 교체일 갱신
        recent_date = f"최근 교체일: {date}"
        next_date = (datetime.strptime(date, "%y%m%d") + timedelta(days=1825)).strftime("%y%m%d")
        next_date_str = f"다음 교체일: {next_date}"
        sheet1["G21"] = f"{recent_date}   {next_date_str}"
        with open(replace_log_filepath, 'a') as replace_log_file:
            replace_log_file.write(f"G21 {recent_date}   {next_date_str}\n")

    # 엑셀 파일 저장
    workbook.save(filepath)

    # 로그 파일 업데이트
    with open(log_filepath, 'a') as log_file:
        log_file.write(f"Date: {year}/{int(month)}/{day} - Updated by: {worker}, Managed by: {manager}\n")

# DB 데이터 가져와서 엑셀 업데이트
async def process_vrs_data():
    await database.connect()
    
    # 모든 시트 데이터를 한번에 가져옵니다.
    queries = {
        "vrs_sheet1": "SELECT * FROM vrs_sheet1",
        "vrs_sheet2": "SELECT * FROM vrs_sheet2",
        "vrs_sheet3": "SELECT * FROM vrs_sheet3"
    }
    data = {sheet_name: await database.fetch_all(query) for sheet_name, query in queries.items()}
    
    # 각 시트 데이터를 병합하여 엑셀 업데이트
    for record in data["vrs_sheet1"]:
        # 같은 date, team, equipment_id에 대한 아이템 데이터를 모아서 전달
        items = {}
        for sheet_name in ["vrs_sheet1", "vrs_sheet2", "vrs_sheet3"]:
            for row in data[sheet_name]:
                if row['date'] == record['date'] and row['team'] == record['team'] and row['equipment_id'] == record['equipment_id']:
                    items[row['item_id']] = row['checked']
        
        update_vrs_excel(record['date'], record['team'], record['worker'], record['manager'], record['equipment_id'], items)
    
    await database.disconnect()

# 메인 함수 실행
if __name__ == "__main__":
    asyncio.run(process_vrs_data())
