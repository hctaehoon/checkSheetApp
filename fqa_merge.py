from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import shutil
from databases import Database
import asyncio
from datetime import datetime, timedelta

# 네트워크 경로로 변경된 DATABASE_URL
# DATABASE_URL = "sqlite:///\\\\110.15.136.98\\data\\test.db"
# local
# 현재 스크립트의 디렉토리 경로를 가져옵니다.
current_dir = os.path.dirname(os.path.abspath(__file__))

# test.db 파일의 전체 경로를 생성합니다.
database_path = os.path.join(current_dir, "test.db")

# DATABASE_URL을 수정합니다.
DATABASE_URL = f"sqlite:///{database_path}"

database = Database(DATABASE_URL)

def is_merged_cell(sheet, row, col):
    """해당 셀이 병합된 셀의 첫 번째 셀인지 확인"""
    cell_coord = f"{get_column_letter(col)}{row}"
    for merged_range in sheet.merged_cells.ranges:
        if cell_coord in merged_range:
            start_coord = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
            return start_coord == cell_coord
    return True  # 병합되지 않은 셀은 True로 처리

# 엑셀 파일 업데이트 및 로그 기록 로직
def update_fqa_excel(date, team, worker, manager, items):
    # 날짜에서 년과 월을 추출
    year = date[:2]  # 예: '240826' -> '24'
    month = date[2:4]  # 예: '240826' -> '08'
    day = date[4:6]  # 예: '240826' -> '26'
    day = int(day)

    # 팀별 폴더 설정
    target_directory = f"./excel/fqa/{team}"
    os.makedirs(target_directory, exist_ok=True)
    
    # 파일 이름 생성 및 경로 설정
    filename = f"FQA_{int(month)}월.xlsx"
    filepath = os.path.join(target_directory, filename)

    # 로그 파일 경로 설정
    log_filename = f"log_{int(month)}월.txt"
    log_filepath = os.path.join(target_directory, log_filename)

    # 교체일 로그 파일 경로 설정
    replace_log_filename = "replace_log.txt"
    replace_log_filepath = os.path.join(target_directory, replace_log_filename)

    # 로그 파일에서 중복 확인
    if os.path.exists(log_filepath):
        with open(log_filepath, 'r') as log_file:
            logs = log_file.readlines()
            for log in logs:
                if f"{year}/{int(month)}/{day}" in log:
                    print(f"이미 기록된 날짜: {year}/{int(month)}/{day}")
                    return  # 이미 기록된 날짜가 있으면 종료

    # 해당 월의 파일이 있는지 확인
    if os.path.exists(filepath):
        workbook = load_workbook(filename=filepath)
    else:
        # 파일이 없으면 기본 템플릿 파일(FQA.xlsx)을 복사하여 새로 생성
        shutil.copyfile("./checksheet/FQA.xlsx", filepath)
        workbook = load_workbook(filename=filepath)

    # 첫 번째 시트 업데이트
    sheet1 = workbook["DMTI-FQA-01-06 (FQA 청소 점검)"]
    sheet1["AB3"] = f"{year}년 {int(month)}월"
    sheet1["F2"] = f"FQA ( {team}조)"
    column_index = 7 + (day - 1)
    rows_to_update_sheet1 = [7, 8, 9, 10, 11, 13, 15, 17, 19, 21, 23, 25, 27, 31]
    for row in rows_to_update_sheet1:
        if is_merged_cell(sheet1, row, column_index):
            sheet1.cell(row=row, column=column_index).value = 'O'

    # 두 번째 시트 업데이트
    sheet2 = workbook["DMTI-FQA-01-06 (청소검증)"]
    sheet2["AB3"] = f"{year}년 {int(month)}월"
    sheet2["F2"] = f"FQA ( {team}조)"
    rows_to_update_sheet2 = [7, 8, 9, 10, 11, 13, 15, 17, 19, 21, 23, 25, 27,29, 31, 33, 35, 37, 39, 41, 43, 45]
    for row in rows_to_update_sheet2:
        if is_merged_cell(sheet2, row, column_index):
            sheet2.cell(row=row, column=column_index).value = 'O'

    # 교체일 로직
    if os.path.exists(replace_log_filepath):
        with open(replace_log_filepath, 'r') as replace_log_file:
            replace_logs = replace_log_file.readlines()
            for line in replace_logs:
                if "G29" in line:
                    sheet1["G29"] = line.strip().replace("G29", "")
                if "G30" in line:
                    sheet1["G30"] = line.strip().replace("G30", "")

    if items.get(11) == 1:  # item_id 11이 1이면 교체일 갱신
        recent_date = f"최근 교체일: {date}"
        next_date = (datetime.strptime(date, "%y%m%d") + timedelta(days=90)).strftime("%y%m%d")
        next_date_str = f"다음 교체일: {next_date}"
        sheet1["G29"] = f"{recent_date}   {next_date_str}"
        with open(replace_log_filepath, 'a') as replace_log_file: 
            replace_log_file.write(f"G29 {recent_date}   {next_date_str}\n")

    if items.get(12) == 1:  # item_id 12가 1이면 교체일 갱신
        recent_date = f"최근 교체일: {date}"
        next_date = (datetime.strptime(date, "%y%m%d") + timedelta(days=365)).strftime("%y%m%d")
        next_date_str = f"다음 교체일: {next_date}"
        sheet1["G30"] = f"{recent_date}   {next_date_str}"
        with open(replace_log_filepath, 'a') as replace_log_file:
            replace_log_file.write(f"G30 {recent_date}   {next_date_str}\n")

    # worker와 manager를 첫 번째 시트의 G33, G34에 입력
    sheet1.cell(row=33, column=column_index).value = worker
    sheet1.cell(row=34, column=column_index).value = manager

    # 두 번째 시트에도 해당 날짜에 맞는 열들에 O 입력
    for row in [33, 34]:
        if is_merged_cell(sheet2, row, column_index):
            sheet2.cell(row=row, column=column_index).value = 'O'

    # 엑셀 파일 저장
    workbook.save(filepath)

    # 로그 파일 업데이트
    with open(log_filepath, 'a') as log_file:
        log_file.write(f"Date: {year}/{int(month)}/{day} - Updated by: {worker}, Managed by: {manager}\n")

# 누락된 날짜를 찾아서 처리
def fill_missing_dates(data, team, worker, manager):
    dates = sorted(set(record['date'] for record in data if record['team'] == team))
    missing_dates = []

    for i in range(len(dates) - 1):
        current_date = datetime.strptime(dates[i], "%y%m%d")
        next_date = datetime.strptime(dates[i + 1], "%y%m%d")
        delta = (next_date - current_date).days
        
        if delta > 1:
            for day in range(1, delta):
                missing_date = (current_date + timedelta(days=day)).strftime("%y%m%d")
                missing_dates.append(missing_date)

    for missing_date in missing_dates:
        print(f"누락된 날짜: {missing_date} - {team} 팀")
        last_valid_record = next(record for record in reversed(data) if record['date'] < missing_date and record['team'] == team)
        update_fqa_excel(missing_date, team, worker, manager, {last_valid_record['item_id']: last_valid_record['checked']})

# DB 데이터 가져와서 엑셀 업데이트
async def process_fqa_data():
    await database.connect()
    query = "SELECT * FROM fqa_sheet1"
    data = await database.fetch_all(query)
    
    for record in data:
        # 같은 date와 team에 대한 아이템 데이터를 모아서 전달
        items = {}
        for row in data:
            if row['date'] == record['date'] and row['team'] == record['team']:
                items[row['item_id']] = row['checked']
        
        update_fqa_excel(record['date'], record['team'], record['worker'], record['manager'], items)
    
    # 누락된 날짜 처리
    teams = set(record['team'] for record in data)
    for team in teams:
        team_data = [record for record in data if record['team'] == team]
        if team_data:
            fill_missing_dates(team_data, team, team_data[0]['worker'], team_data[0]['manager'])
    
    await database.disconnect()

# 메인 함수 실행
if __name__ == "__main__":
    asyncio.run(process_fqa_data())
